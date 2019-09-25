from openpyxl import *


class excel():
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        self.ws = self.wb['Sheet1']

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column,ignoreRow):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            if i == ignoreRow:
                continue
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

    # 设置某个单元格的值
    def setCellsValue(self, beginRow, list):
        try:
            for i, companyInfo in enumerate(list):
                if not companyInfo:
                    continue
                row = beginRow + i + 2
                self.ws.cell(row=row, column=4).value = companyInfo['统一社会信用代码']
                self.ws.cell(row=row, column=5).value = companyInfo['组织机构代码']
                self.ws.cell(row=row, column=6).value = companyInfo['注册号']
                self.ws.cell(row=row, column=7).value = companyInfo['经营状态']
                self.ws.cell(row=row, column=8).value = companyInfo['公司类型']
                self.ws.cell(row=row, column=9).value = companyInfo['成立日期']
                self.ws.cell(row=row, column=10).value = companyInfo['法定代表人']
                self.ws.cell(row=row, column=11).value = companyInfo['营业期限']
                self.ws.cell(row=row, column=12, ).value = companyInfo['注册资本']
                self.ws.cell(row=row, column=13).value = companyInfo['发照日期']
                self.ws.cell(row=row, column=14).value = companyInfo['登记机关']
                self.ws.cell(row=row, column=15).value = companyInfo['企业地址']
                self.ws.cell(row=row, column=16).value = companyInfo['经营范围']
                # self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=i + 2, column=4).value = "writefail"
            self.wb.save(self.file)

    # 设置某个单元格的值(百度信用查询使用)
    def setCellsValue4BaiduXinyong(self, beginRow, list):
        try:
            for i, companyInfo in enumerate(list):
                if not companyInfo:
                    continue
                row = beginRow + i + 2
                self.ws.cell(row=row, column=4).value = companyInfo['统一社会信用代码']
                self.ws.cell(row=row, column=5).value = companyInfo['组织机构代码']
                self.ws.cell(row=row, column=6).value = companyInfo['工商注册号']
                self.ws.cell(row=row, column=7).value = companyInfo['经营状态']
                self.ws.cell(row=row, column=8).value = companyInfo['企业类型']
                self.ws.cell(row=row, column=9).value = companyInfo['成立日期']
                self.ws.cell(row=row, column=10).value = companyInfo['法定代表人']
                self.ws.cell(row=row, column=11).value = companyInfo['营业期限']
                self.ws.cell(row=row, column=12).value = companyInfo['注册资本']
                self.ws.cell(row=row, column=13).value = companyInfo['注册地址']
                self.ws.cell(row=row, column=14).value = companyInfo['登记机关']
                self.ws.cell(row=row, column=15).value = companyInfo['经营范围']
            self.wb.save(self.file)
        except Exception as e:
            print(e)
            self.ws.cell(row=i + 2, column=4).value = "writefail"
            self.wb.save(self.file)

def writeData(excel,list):
    for i,companyInfo in enumerate(list):
        if not companyInfo:
            continue
        excel.setCellValue(i+2, 4, companyInfo['统一社会信用代码'])
        excel.setCellValue(i+2, 5, companyInfo['组织机构代码'])
        excel.setCellValue(i+2, 6, companyInfo['注册号'])
        excel.setCellValue(i+2, 7, companyInfo['经营状态'])
        excel.setCellValue(i+2, 8, companyInfo['公司类型'])
        excel.setCellValue(i+2, 9, companyInfo['成立日期'])
        excel.setCellValue(i+2, 10, companyInfo['法定代表人'])
        excel.setCellValue(i+2, 11, companyInfo['营业期限'])
        excel.setCellValue(i+2, 12, companyInfo['注册资本'])
        excel.setCellValue(i+2, 13, companyInfo['发照日期'])
        excel.setCellValue(i+2, 14, companyInfo['登记机关'])
        excel.setCellValue(i+2, 15, companyInfo['企业地址'])
        excel.setCellValue(i+2, 16, companyInfo['经营范围'])
# if __name__ == '__main__':
#     excel = excel("C:/Users/xielw/Desktop/公司处理/测试.xlsx")
#     companyNames = excel.getColValues(1, 1)
#
#     list = []
#     for i, companyName in enumerate(companyNames):
#         companyInfo = {}
#         # companyInfo['公司名称'] = companyName  # 公司名称
#         companyInfo['统一社会信用代码'] ='统一社会信用代码' + str(i)  # 统一社会信用代码
#         companyInfo['组织机构代码'] = '组织机构代码' + str(i)  # 组织机构代码
#         companyInfo['工商注册号'] = '工商注册号' + str(i)  # 工商注册号
#         companyInfo['经营状态'] = '经营状态' + str(i)  # 经营状态
#         companyInfo['企业类型'] = '企业类型' + str(i)  # 企业类型
#         companyInfo['成立日期'] = '成立日期' + str(i)  # 成立日期
#         companyInfo['法定代表人'] = '法定代表人' + str(i)  # 法定代表人
#         companyInfo['营业期限'] = '营业期限' + str(i)  # 营业期限
#         companyInfo['注册资本'] = '注册资本' + str(i)  # 注册资本
#         companyInfo['注册地址'] = '注册地址' + str(i)  # 注册地址
#         companyInfo['登记机关'] = '登记机关' + str(i)  # 登记机关
#         companyInfo['经营范围'] = '经营范围' + str(i)  # 经营范围
#         list.append(companyInfo)
# print(list)
# excel.setCellsValue4BaiduXinyong(0,list)