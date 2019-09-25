import requests
from openpyxl import load_workbook
from selenium import webdriver
import re
import urllib

from company.userAgent import getUserAgent2


def get_company():
    wb=load_workbook('企查查-11480.xlsx')
    ws = wb['Sheet1']
    # cell_range = ws['A2:A11481']
    cell_range = ws['A4:A10']

    companys=[]

    for i in cell_range:
        companys.append(i[0].value)
    return companys

def get_url(companys):
    driver = webdriver.Chrome()
    driver.get('https://www.company_xin_baidu.com/')
    hrefs = []
    for i in companys:
        driver.find_element_by_xpath('//input[@class="index-searchkey form-control input-lg"]').send_keys(i)
        driver.find_element_by_xpath('//input[@id="V3_Search_bt"]').click()
        company = driver.find_element_by_xpath('//a[@class="ma_h1"]')
        href = company.get_attribute('href')
        hrefs.append(href)
        driver.get('https://www.company_xin_baidu.com/')
    return hrefs

def ger_result(hrefs):
    headers = ("User-Agent",
               "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.22 Safari/537.36 SE 2.X MetaSr 1.0")
    opener = urllib.request.build_opener()
    opener.addheaders = [headers]
    urllib.request.install_opener(opener)
    # url = 'http://www.baidu.com'
    for url in hrefs:
        data = urllib.request.urlopen(url).read().decode('utf-8')
        company={}
        person_path = '<h2 class="seo font-20">(.*?)</h2>'
        person_name = re.compile(person_path).findall(data)
        register_path ='注册资本 </td> <td width="30%">\s*(\d+[\u0391-\uFFE5]+)\s*</td>'
        register__money = re.compile(register_path).findall(data)

        really_path ='实缴资本 </td> <td width="30%">\s*(\d+[\u0391-\uFFE5]+)\s*</td>'
        really__money = re.compile(register_path).findall(data)

        status_path ='经营状态</td> <td class="">\s*(.*?)\s*</td>'
        status = re.compile(status_path).findall(data)

        publish_path ='成立日期</td> <td class="">\s*(.*?)\s*</td>'
        publish_data =re.compile(publish_path).findall(data)

        credit_pa = '<td class="tb">统一社会信用代码</td> <td class="">\s*(.*)\s*</td>'
        credit_code = re.compile(credit_pa).findall(data)

        tax_pa = '纳税人识别号</td> <td class="">\s*(.*?)\s*</td>'
        tax_num = re.compile(credit_pa).findall(data)

        register_path ='注册号</td> <td class="">\s*(.*?)\s*</td>'
        register__num = re.compile(really_path).findall(data)
        organ_path ='组织机构代码</td> <td class="">\s*(.*?)\s*</td>'
        organ__num = re.compile(organ_path).findall(data)

        type_path ='企业类型</td> <td class="">\s*(.*?)\s*</td>'
        type = re.compile(type_path).findall(data)

        industry_path ='所属行业</td> <td class="">\s*(.*?)\s*</td>'
        industry = re.compile(industry_path).findall(data)

        hezhun_path ='核准日期</td> <td class="" style="max-width:301px;">\s*(.*?)\s*</td>'

        hezhun_date =re.compile(hezhun_path).findall(data)

        dengjijiguan_path ='登记机关</td> <td class="">\s*(.*?)\s*</td>'

        dengjijiguan = re.compile(dengjijiguan_path).findall(data)

        diqu_path ='所属地区</td> <td class="" style="max-width:301px;">\s*(.*?)\s*</td>'

        diqu = re.compile(diqu_path).findall(data)

        english_path ='英文名</td> <td class="">\s*(.*?)\s*</td>'

        englishi_name = re.compile(english_path).findall(data)

        before_name_path ='曾用名\s*</td> <td class="">\s*(.*?)\s*</td>'

        before_name = re.compile(before_name_path).findall(data)

        insure_path ='参保人数\s*</td> <td class="">\s*(.*?)\s*</td>'

        insure_num = re.compile(insure_path).findall(data)

        scale_path ='人员规模\s*</td> <td class="">\s*(.*?)\s*</td>'

        scale = re.compile(insure_path).findall(data)

        work_time_path = '营业期限\s*</td> <td class="">\s*(.*?)\s*</td>'

        work_time = re.compile(work_time_path).findall(data)

        address_apth ='企业地址</td> <td class="" colspan="3">\s*(.*)\s*<'

        adress = re.compile(address_apth).findall(data)

        jingyingfanwei_path = '经营范围</td> <td class="" colspan="3">\s*(.*?)\s*</td>'

        jingyingfanwei = re.compile(jingyingfanwei_path).findall(data)

        print(person_name)
        print(credit_code)
        print(register__money)
        print(really__money)
        print(status)
        print(publish_data)
        print(tax_num)
        print(register__num)
        print(organ__num)
        print(type)
        print(industry)
        print(hezhun_date, dengjijiguan, diqu, englishi_name, before_name, insure_num, scale, work_time, adress,
              jingyingfanwei)


def parData(driver, companyName, proxyIp):
    company = {}

    ## 获取url
    driver.get('https://www.company_xin_baidu.com/')
    driver.find_element_by_xpath('//input[@class="index-searchkey form-control input-lg"]').send_keys(companyName)
    driver.find_element_by_xpath('//input[@id="V3_Search_bt"]').click()
    company = driver.find_element_by_xpath('//a[@class="ma_h1"]')
    url = company.get_attribute('href')
    driver.get('https://www.company_xin_baidu.com/')

    # 爬出数据
    uerAgent = getUserAgent2()
    headers = ("User-Agent", uerAgent)

    httpproxy_handler = urllib.request.ProxyHandler({"http": "http://" + proxyIp, "https": "http://" + proxyIp})
    opener = urllib.request.build_opener(httpproxy_handler)
    opener.addheaders = [headers]

    urllib.request.install_opener(opener)

    data = urllib.request.urlopen(url).read().decode('utf-8')
    person_path = '<h2 class="seo font-20">(.*?)</h2>'
    person_name = re.compile(person_path).findall(data)
    register_path = '注册资本 </td> <td width="30%">\s*(\d+[\u0391-\uFFE5]+)\s*</td>'
    register__money = re.compile(register_path).findall(data)

    really_path = '实缴资本 </td> <td width="30%">\s*(\d+[\u0391-\uFFE5]+)\s*</td>'
    really__money = re.compile(register_path).findall(data)

    status_path = '经营状态</td> <td class="">\s*(.*?)\s*</td>'
    status = re.compile(status_path).findall(data)

    publish_path = '成立日期</td> <td class="">\s*(.*?)\s*</td>'
    publish_data = re.compile(publish_path).findall(data)

    credit_pa = '<td class="tb">统一社会信用代码</td> <td class="">\s*(.*)\s*</td>'
    credit_code = re.compile(credit_pa).findall(data)

    tax_pa = '纳税人识别号</td> <td class="">\s*(.*?)\s*</td>'
    tax_num = re.compile(credit_pa).findall(data)

    register_path = '注册号</td> <td class="">\s*(.*?)\s*</td>'
    register__num = re.compile(really_path).findall(data)
    organ_path = '组织机构代码</td> <td class="">\s*(.*?)\s*</td>'
    organ__num = re.compile(organ_path).findall(data)

    type_path = '企业类型</td> <td class="">\s*(.*?)\s*</td>'
    type = re.compile(type_path).findall(data)

    industry_path = '所属行业</td> <td class="">\s*(.*?)\s*</td>'
    industry = re.compile(industry_path).findall(data)

    hezhun_path = '核准日期</td> <td class="" style="max-width:301px;">\s*(.*?)\s*</td>'

    hezhun_date = re.compile(hezhun_path).findall(data)

    dengjijiguan_path = '登记机关</td> <td class="">\s*(.*?)\s*</td>'

    dengjijiguan = re.compile(dengjijiguan_path).findall(data)

    diqu_path = '所属地区</td> <td class="" style="max-width:301px;">\s*(.*?)\s*</td>'

    diqu = re.compile(diqu_path).findall(data)

    english_path = '英文名</td> <td class="">\s*(.*?)\s*</td>'

    englishi_name = re.compile(english_path).findall(data)

    before_name_path = '曾用名\s*</td> <td class="">\s*(.*?)\s*</td>'

    before_name = re.compile(before_name_path).findall(data)

    insure_path = '参保人数\s*</td> <td class="">\s*(.*?)\s*</td>'

    insure_num = re.compile(insure_path).findall(data)

    scale_path = '人员规模\s*</td> <td class="">\s*(.*?)\s*</td>'

    scale = re.compile(insure_path).findall(data)

    work_time_path = '营业期限\s*</td> <td class="">\s*(.*?)\s*</td>'

    work_time = re.compile(work_time_path).findall(data)

    address_apth = '企业地址</td> <td class="" colspan="3">\s*(.*)\s*<'

    adress = re.compile(address_apth).findall(data)

    jingyingfanwei_path = '经营范围</td> <td class="" colspan="3">\s*(.*?)\s*</td>'

    jingyingfanwei = re.compile(jingyingfanwei_path).findall(data)

    print(hezhun_date, dengjijiguan, diqu, englishi_name, before_name, insure_num, scale, work_time, adress,
          jingyingfanwei)

if __name__ == '__main__':
    ger_result(get_url(get_company()))
